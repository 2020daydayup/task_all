import openpyxl

class BtcAnalysis():
    "处理比特币excel类"
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.load_workbook(self.filename)

    def get_data(self):
        "读取sheet"
        data = self.wb['btc']
        for i in range(1,data.max_row):
            date = data[f'a{i+1}'].value
            year = date.split('-')[0]

            try:
                ws = self.wb[year]
                ws.append([date, data[f'b{i+1}'].value])
            except:
                ws = self.wb.create_sheet(title=year)
                ws.append(['日期', '时间'])
                ws.append([date, data[f'b{i+1}'].value])
        self.wb.save(f'test_{self.filename}')

    def add_average(self):
        "插入平均值"
        sheets = self.wb.sheetnames
        for sheet in  sheets:
            self.wb[sheet].append(['平均值',f'=AVERAGE(b2:b{self.wb[sheet].max_row})'])
            self.wb.save(f'test_{self.filename}')



if __name__ == '__main__':
    btc = BtcAnalysis("btc.xlsx")
    btc.get_data()
    btc.add_average()
